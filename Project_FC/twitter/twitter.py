import time
from selenium import webdriver
from openpyxl import load_workbook
from konlpy.tag import Okt

URL = "https://twitter.com/search?q={}&src=typd"
DRIVER_DIR = '/Users/temp/Project_FC/chromedriver'
SAVE_DIR = 'test.xlsx'

def twitter(keyword):
    try:
        driver = webdriver.Chrome(DRIVER_DIR)
        driver.implicitly_wait(10) # 암묵적으로 웹 자원을 (최대) 10초 기다리기
        driver.get(URL.format(str(keyword)))
        
        no_page = 0
        while no_page < 10:
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
            no_page += 1
            time.sleep(1.5)
        # 모든 트윗 리스트
        content = driver.find_elements_by_css_selector('div.content')
        print("content-length: ", len(content))
        
        result = []
        # 반복문으로 각각의 트윗 뽑아오기
        for i in content:
            # 트윗 텍스트
            cont = i.find_element_by_css_selector('p.tweet-text')
            # 트윗 업로드 시간
            timestamp = i.find_element_by_css_selector('a.tweet-timestamp')
            result.append([cont.text.strip(), timestamp.get_attribute("title")])
            # result = [[텍스트, 시간], [텍스트, 시간], [텍스트, 시간], ...]
            print(cont.text.strip(), timestamp.get_attribute("title"))
            print('------------------------------------------')
        save_excel(result)
    except Exception as e:
        print(e)
    finally:
        driver.quit()  

def save_excel(result):
    try:
        wb = load_workbook(SAVE_DIR)  
        ws = wb.create_sheet(title='twitter')
        for idx, re in enumerate(result):
            ws['A' + str(idx + 1)] = re[0]
            ws['B' + str(idx + 1)] = re[1]
        wb.save(SAVE_DIR)
    finally:
        wb.close()

def get_excel():
    result = []
    try:
        wb = load_workbook(SAVE_DIR, read_only=True)
        ws = wb['twitter']
        for row in ws.rows:
            result.append(row[0].value)
            # for cell in row: 
                # print(cell.value)
    finally:
        wb.close()
    return result
    
def get_content(result):
    ok = Okt()
    content = {} # 실제 값을 표현할
    for re in result:
        temp = ok.pos(re) # 값 -> ('단어', '품사'), ('단어', '품사'), ('단어', '품사'), ...
        for t in temp:
            if t[1] == 'Hashtag': # 해시태그?
                if not (t[0] in content): # 이미 결과 값이 초기화 되어있는지?
                    content[t[0]] = 0 # 없다면 초기화
                content[t[0]] += 1 # 1을 더하는 연산
    content = sorted(content.items(), key = lambda x:x[1], reverse = True)

    for k,v in content:
        print("({}){}".format(k, v), end = ' ')

if __name__ == "__main__":
    keyword = input('keyword?')
    twitter(keyword)
    # get_content(get_excel())
